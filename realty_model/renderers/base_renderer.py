from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from realty_model.config import PALETTE


class BaseRenderer:
    HDR_BG    = PALETTE['navy']
    HDR_FG    = PALETTE['white']
    SUB_BG    = PALETTE['steel']
    SEC_BG    = PALETTE['light_blue']
    GREEN_FG  = PALETTE['green']
    GREEN_BG  = PALETTE['light_green']
    RED_FG    = PALETTE['red']
    RED_BG    = PALETTE['light_red']
    GOLD      = PALETTE['gold']
    GOLD_BG   = PALETTE['light_gold']
    GRAY_BG   = PALETTE['gray_bg']
    DARK_GRAY = PALETTE['dark_gray']

    def _fill(self, color: str):
        return PatternFill('solid', fgColor=color)

    def _font(self, bold=False, color=PALETTE['black'], size=10, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    def _align(self, h='left', v='center', wrap=False, indent=0):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap, indent=indent)

    def _border(self, color=PALETTE['border']):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def style_header(self, cell, text=None):
        if text is not None:
            cell.value = text
        cell.font = self._font(bold=True, color=self.HDR_FG, size=11)
        cell.fill = self._fill(self.HDR_BG)
        cell.alignment = self._align(h='center', wrap=True)

    def style_subheader(self, cell, text=None):
        if text is not None:
            cell.value = text
        cell.font = self._font(bold=True, color=self.HDR_FG, size=10)
        cell.fill = self._fill(self.SUB_BG)
        cell.alignment = self._align(h='center')

    def style_section(self, cell, text=None):
        if text is not None:
            cell.value = text
        cell.font = self._font(bold=True, size=10)
        cell.fill = self._fill(self.SEC_BG)
        cell.alignment = self._align(h='left')

    def style_label(self, cell, text=None, indent=0):
        if text is not None:
            cell.value = text
        cell.font = self._font(size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=indent)

    def style_total(self, cell, text=None):
        if text is not None:
            cell.value = text
        cell.font = self._font(bold=True, size=10)
        cell.fill = self._fill(self.DARK_GRAY)
        cell.alignment = self._align(h='right')

    def fmt_currency(self, cell):
        cell.number_format = '_($* #,##0_);_($* (#,##0);_($* "-"_);_(@_)'

    def fmt_percent(self, cell, decimals=1):
        cell.number_format = f'0.{"0" * decimals}%'

    def fmt_number(self, cell, decimals=2):
        cell.number_format = f'#,##0.{"0" * decimals}'

    def fmt_integer(self, cell):
        cell.number_format = '#,##0'

    def set_col_width(self, ws, col: int, width: float):
        ws.column_dimensions[get_column_letter(col)].width = width

    def apply_border_range(self, ws, min_row, max_row, min_col, max_col):
        b = self._border()
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = b

    def write_value_cell(self, ws, row, col, value, fmt=None, bold=False,
                         bg=None, fg=None, align_h='right'):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = self._font(bold=bold, color=fg or PALETTE['black'])
        cell.alignment = self._align(h=align_h)
        if bg:
            cell.fill = self._fill(bg)
        if fmt == 'currency':
            self.fmt_currency(cell)
        elif fmt == 'percent':
            self.fmt_percent(cell)
        elif fmt == 'percent2':
            self.fmt_percent(cell, decimals=2)
        elif fmt == 'number':
            self.fmt_number(cell)
        elif fmt == 'integer':
            self.fmt_integer(cell)
        return cell

    def add_title_row(self, ws, row, title, cols_span, bg=None, size=14):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols_span)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = Font(bold=True, color=self.HDR_FG, size=size)
        cell.fill = self._fill(bg or self.HDR_BG)
        cell.alignment = self._align(h='center')
        ws.row_dimensions[row].height = 30
