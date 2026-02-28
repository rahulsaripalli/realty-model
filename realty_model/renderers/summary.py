from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from realty_model.config import PALETTE, IRR_THRESHOLDS, DSCR_THRESHOLDS, COC_THRESHOLDS
from realty_model.renderers.base_renderer import BaseRenderer
from realty_model.models.results import FinancialResults
from realty_model.models.property import ProFormaInput


def _irr_color(irr: float) -> str:
    if irr != irr:  # nan
        return PALETTE['gray_bg']
    if irr >= IRR_THRESHOLDS['strong']:
        return PALETTE['light_green']
    if irr >= IRR_THRESHOLDS['okay']:
        return PALETTE['light_gold']
    return PALETTE['light_red']


def _dscr_color(dscr: float) -> str:
    if dscr >= DSCR_THRESHOLDS['strong']:
        return PALETTE['light_green']
    if dscr >= DSCR_THRESHOLDS['okay']:
        return PALETTE['light_gold']
    return PALETTE['light_red']


def _coc_color(coc: float) -> str:
    if coc >= COC_THRESHOLDS['strong']:
        return PALETTE['light_green']
    if coc >= COC_THRESHOLDS['okay']:
        return PALETTE['light_gold']
    return PALETTE['light_red']


class SummaryRenderer(BaseRenderer):
    def render(self, wb, results: FinancialResults, inp: ProFormaInput):
        ws = wb.create_sheet("Summary Dashboard")
        ws.sheet_view.showGridLines = False

        m = results.key_metrics
        addr = inp.property_info.address

        # Title
        self.add_title_row(ws, 1, f"Investment Analysis  |  {addr}", 10)
        ws.row_dimensions[1].height = 35

        # Column widths
        for col, w in [(1, 28), (2, 18), (3, 2), (4, 28), (5, 18), (6, 2), (7, 28), (8, 18)]:
            self.set_col_width(ws, col, w)

        # --- Property Snapshot (col 1-2) ---
        row = 3
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        self.style_subheader(ws.cell(row=row, column=1), "PROPERTY SNAPSHOT")
        row += 1

        snapshot = [
            ("Property Type",   inp.property_info.property_type,   None),
            ("Year Built",      inp.property_info.year_built,       None),
            ("Units",           inp.property_info.units,            'integer'),
            ("Rentable SF",     inp.property_info.square_feet,      'integer'),
            ("Purchase Price",  m.purchase_price,                    'currency'),
            ("Loan Amount",     m.loan_amount,                       'currency'),
            ("Equity Invested", m.equity_invested,                   'currency'),
            ("LTV",             m.ltv,                               'percent'),
            ("Interest Rate",   inp.financing.interest_rate,         'percent2'),
            ("Loan Term",       f"{inp.financing.loan_term_years} yrs", None),
        ]

        for label, value, fmt in snapshot:
            bg = self.GRAY_BG if row % 2 == 0 else None
            cell_l = ws.cell(row=row, column=1, value=label)
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')
            if bg:
                cell_l.fill = self._fill(bg)

            cell_v = ws.cell(row=row, column=2, value=value)
            cell_v.font = self._font(bold=True, size=10)
            cell_v.alignment = self._align(h='right')
            if bg:
                cell_v.fill = self._fill(bg)
            if fmt == 'currency':
                self.fmt_currency(cell_v)
            elif fmt == 'percent':
                self.fmt_percent(cell_v)
            elif fmt == 'percent2':
                self.fmt_percent(cell_v, decimals=2)
            elif fmt == 'integer':
                self.fmt_integer(cell_v)
            row += 1

        # --- Income Summary (col 1-2 continued) ---
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        self.style_subheader(ws.cell(row=row, column=1), "INCOME SUMMARY (YEAR 1)")
        row += 1

        income_items = [
            ("Gross Scheduled Rent", inp.income.gross_scheduled_rent, 'currency'),
            ("Vacancy Rate",         inp.income.vacancy_rate,          'percent'),
            ("Other Income",         inp.income.other_income,           'currency'),
            ("Rent Growth Rate",     inp.income.rent_growth_rate,       'percent'),
        ]
        for label, value, fmt in income_items:
            bg = self.GRAY_BG if row % 2 == 0 else None
            cell_l = ws.cell(row=row, column=1, value=label)
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')
            if bg:
                cell_l.fill = self._fill(bg)
            cell_v = ws.cell(row=row, column=2, value=value)
            cell_v.font = self._font(bold=True, size=10)
            cell_v.alignment = self._align(h='right')
            if bg:
                cell_v.fill = self._fill(bg)
            if fmt == 'currency':
                self.fmt_currency(cell_v)
            elif fmt == 'percent':
                self.fmt_percent(cell_v)
            row += 1

        # --- Key Metrics (col 4-5) ---
        km_row = 3
        ws.merge_cells(start_row=km_row, start_column=4, end_row=km_row, end_column=5)
        self.style_subheader(ws.cell(row=km_row, column=4), "KEY METRICS")
        km_row += 1

        key_metrics = [
            ("Going-In Cap Rate",      m.going_in_cap_rate,             'percent2', None),
            ("Gross Rent Multiplier",   m.grm,                           'number',   None),
            ("DSCR (Year 1)",          m.dscr_year1,                    'number',   _dscr_color(m.dscr_year1)),
            ("Break-Even Occupancy",   m.break_even_occupancy,          'percent',  None),
            ("Cash-on-Cash (Yr 1)",    m.cash_on_cash_year1,            'percent2', _coc_color(m.cash_on_cash_year1)),
            ("Cash-on-Cash (Stab.)",   m.cash_on_cash_stabilized,       'percent2', _coc_color(m.cash_on_cash_stabilized)),
            ("Levered IRR",            m.levered_irr,                   'percent2', _irr_color(m.levered_irr)),
            ("Unlevered IRR",          m.unlevered_irr,                 'percent2', _irr_color(m.unlevered_irr)),
            ("Equity Multiple",        m.equity_multiple,               'number',   None),
            ("NPV",                    m.npv,                           'currency', None),
            ("Exit Value",             m.exit_value,                    'currency', None),
            ("Net Exit Proceeds",      m.exit_proceeds,                 'currency', None),
        ]

        for label, value, fmt, color in key_metrics:
            bg_l = self.GRAY_BG if km_row % 2 == 0 else None
            cell_l = ws.cell(row=km_row, column=4, value=label)
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')
            if bg_l:
                cell_l.fill = self._fill(bg_l)

            display_val = value
            if fmt in ('percent', 'percent2') and value != value:
                display_val = 'N/A'
                fmt = None

            cell_v = ws.cell(row=km_row, column=5, value=display_val)
            cell_v.font = self._font(bold=True, size=10)
            cell_v.alignment = self._align(h='right')
            if color:
                cell_v.fill = self._fill(color)
                cell_l.fill = self._fill(color)
            elif bg_l:
                cell_v.fill = self._fill(bg_l)

            if fmt == 'currency':
                self.fmt_currency(cell_v)
            elif fmt == 'percent':
                self.fmt_percent(cell_v)
            elif fmt == 'percent2':
                self.fmt_percent(cell_v, decimals=2)
            elif fmt == 'number':
                self.fmt_number(cell_v)
            km_row += 1

        # --- Hold Assumptions (col 7-8) ---
        ha_row = 3
        ws.merge_cells(start_row=ha_row, start_column=7, end_row=ha_row, end_column=8)
        self.style_subheader(ws.cell(row=ha_row, column=7), "HOLD ASSUMPTIONS")
        ha_row += 1

        hold_items = [
            ("Hold Period",        f"{inp.assumptions.hold_period_years} years",  None),
            ("Discount Rate",      inp.assumptions.discount_rate,                 'percent'),
            ("Exit Cap Rate",      inp.assumptions.exit_cap_rate,                 'percent2'),
            ("Cost of Sale",       inp.assumptions.cost_of_sale,                  'percent'),
            ("Expense Growth",     inp.expenses.expense_growth_rate,              'percent'),
        ]
        for label, value, fmt in hold_items:
            bg = self.GRAY_BG if ha_row % 2 == 0 else None
            cell_l = ws.cell(row=ha_row, column=7, value=label)
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')
            if bg:
                cell_l.fill = self._fill(bg)
            cell_v = ws.cell(row=ha_row, column=8, value=value)
            cell_v.font = self._font(bold=True, size=10)
            cell_v.alignment = self._align(h='right')
            if bg:
                cell_v.fill = self._fill(bg)
            if fmt == 'percent':
                self.fmt_percent(cell_v)
            elif fmt == 'percent2':
                self.fmt_percent(cell_v, decimals=2)
            ha_row += 1

        # --- Chart data (hidden columns) ---
        # Write NOI and cash flow data for chart into cols 10-12
        cf_df = results.annual_cash_flows
        chart_start = 2
        ws.cell(row=chart_start, column=10, value='Year')
        ws.cell(row=chart_start, column=11, value='NOI')
        ws.cell(row=chart_start, column=12, value='Cash Flow')
        for i, row_data in cf_df.iterrows():
            r = chart_start + 1 + i
            ws.cell(row=r, column=10, value=row_data['Year'])
            ws.cell(row=r, column=11, value=row_data['NOI'])
            ws.cell(row=r, column=12, value=row_data['Net Cash Flow'])
        # Hide chart data columns
        for col in [10, 11, 12]:
            ws.column_dimensions[get_column_letter(col)].hidden = True

        n_years = len(cf_df)
        chart_data_min = chart_start + 1
        chart_data_max = chart_start + n_years

        # Line chart: NOI and Cash Flow
        lc = LineChart()
        lc.title = "10-Year NOI & Cash Flow"
        lc.style = 10
        lc.y_axis.title = "Dollars ($)"
        lc.x_axis.title = "Year"
        lc.width = 18
        lc.height = 12

        noi_ref = Reference(ws, min_col=11, min_row=chart_start, max_row=chart_data_max)
        cf_ref = Reference(ws, min_col=12, min_row=chart_start, max_row=chart_data_max)
        lc.add_data(noi_ref, titles_from_data=True)
        lc.add_data(cf_ref, titles_from_data=True)
        cats = Reference(ws, min_col=10, min_row=chart_data_min, max_row=chart_data_max)
        lc.set_categories(cats)
        ws.add_chart(lc, "A22")

        # Bar chart: EGI vs Expenses
        ws.cell(row=chart_start, column=14, value='Year')
        ws.cell(row=chart_start, column=15, value='EGI')
        ws.cell(row=chart_start, column=16, value='Expenses')
        for i, row_data in cf_df.iterrows():
            r = chart_start + 1 + i
            ws.cell(row=r, column=14, value=row_data['Year'])
            ws.cell(row=r, column=15, value=row_data['EGI'])
            ws.cell(row=r, column=16, value=row_data['Total Expenses'])
        for col in [14, 15, 16]:
            ws.column_dimensions[get_column_letter(col)].hidden = True

        bc = BarChart()
        bc.type = "col"
        bc.title = "EGI vs. Operating Expenses"
        bc.style = 10
        bc.y_axis.title = "Dollars ($)"
        bc.x_axis.title = "Year"
        bc.width = 18
        bc.height = 12

        egi_ref = Reference(ws, min_col=15, min_row=chart_start, max_row=chart_data_max)
        exp_ref = Reference(ws, min_col=16, min_row=chart_start, max_row=chart_data_max)
        bc.add_data(egi_ref, titles_from_data=True)
        bc.add_data(exp_ref, titles_from_data=True)
        cats2 = Reference(ws, min_col=14, min_row=chart_data_min, max_row=chart_data_max)
        bc.set_categories(cats2)
        ws.add_chart(bc, "E22")
