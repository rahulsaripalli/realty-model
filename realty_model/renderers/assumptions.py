"""
Assumptions sheet — the user-editable control panel.
All other formula-driven sheets reference cells here.
"""
from openpyxl.styles import Font
from realty_model.renderers.base_renderer import BaseRenderer
from realty_model.renderers.constants import (
    ASM_PURCHASE_PRICE_ROW, ASM_LOAN_AMOUNT_ROW, ASM_INTEREST_RATE_ROW,
    ASM_LOAN_TERM_ROW, ASM_IO_YEARS_ROW, ASM_CLOSING_COSTS_ROW,
    ASM_EQUITY_ROW, ASM_LTV_ROW,
    ASM_GROSS_RENT_ROW, ASM_VACANCY_ROW, ASM_OTHER_INCOME_ROW, ASM_RENT_GROWTH_ROW,
    ASM_TAXES_ROW, ASM_INSURANCE_ROW, ASM_MGMT_ROW, ASM_MAINTENANCE_ROW,
    ASM_UTILITIES_ROW, ASM_LANDSCAPING_ROW, ASM_CAPEX_ROW, ASM_EXP_GROWTH_ROW,
    ASM_HOLD_YEARS_ROW, ASM_DISCOUNT_RATE_ROW, ASM_EXIT_CAP_ROW, ASM_COST_OF_SALE_ROW,
)
from realty_model.models.property import ProFormaInput


class AssumptionsRenderer(BaseRenderer):
    def render(self, wb, inp: ProFormaInput):
        ws = wb.create_sheet("Assumptions")
        ws.sheet_view.showGridLines = False

        self.add_title_row(ws, 1, "⚙  Model Assumptions  —  Edit yellow cells to update the model", 4)
        ws.row_dimensions[1].height = 30

        # Column widths
        self.set_col_width(ws, 1, 30)
        self.set_col_width(ws, 2, 18)
        self.set_col_width(ws, 3, 6)
        self.set_col_width(ws, 4, 30)

        p = inp.property_info
        f = inp.financing
        i = inp.income
        e = inp.expenses
        a = inp.assumptions
        wf = inp.waterfall

        EDIT_BG = 'FFF2CC'   # yellow = user-editable
        CALC_BG = 'E2EFDA'   # green  = calculated (do not edit)

        def section(row, label):
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            self.style_subheader(ws.cell(row=row, column=1), label)

        def input_row(row, label, value, fmt='currency', editable=True, note=''):
            bg = EDIT_BG if editable else CALC_BG
            cell_l = ws.cell(row=row, column=1, value=label)
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')

            cell_v = ws.cell(row=row, column=2, value=value)
            cell_v.font = self._font(bold=True, size=10)
            cell_v.fill = self._fill(bg)
            cell_v.alignment = self._align(h='right')

            if fmt == 'currency':   self.fmt_currency(cell_v)
            elif fmt == 'pct':      self.fmt_percent(cell_v, 2)
            elif fmt == 'int':      self.fmt_integer(cell_v)
            elif fmt == 'text':     pass

            if note:
                cell_n = ws.cell(row=row, column=4, value=note)
                cell_n.font = self._font(size=9, italic=True, color='888888')

            # Alternate row shading
            alt_bg = 'F7F7F7' if row % 2 == 0 else None
            if alt_bg and editable:
                pass  # keep yellow for editables
            elif alt_bg:
                cell_l.fill = self._fill(alt_bg)
                cell_v.fill = self._fill(CALC_BG)

        # ── PROPERTY ────────────────────────────────────────────────────────
        section(2, "PROPERTY")
        input_row(3, "Address",          p.address,                    fmt='text',    note='')
        input_row(ASM_PURCHASE_PRICE_ROW, "Purchase Price",  p.purchase_price,         fmt='currency')
        input_row(5,  "Property Type",   p.property_type,              fmt='text')
        input_row(6,  "Year Built",      p.year_built,                 fmt='int')
        input_row(7,  "Units",           p.units,                      fmt='int')
        input_row(8,  "Square Feet",     p.square_feet,                fmt='int')

        # ── FINANCING ────────────────────────────────────────────────────────
        section(10, "FINANCING")
        input_row(11, "Interest Rate",   f.interest_rate,              fmt='pct',     note='Annual (e.g. 0.0675 = 6.75%)')
        input_row(ASM_LOAN_AMOUNT_ROW,   "Loan Amount",    f.loan_amount,            fmt='currency')
        input_row(ASM_INTEREST_RATE_ROW, "Interest Rate",  f.interest_rate,          fmt='pct')
        input_row(ASM_LOAN_TERM_ROW,     "Loan Term (yrs)",f.loan_term_years,        fmt='int')
        input_row(ASM_IO_YEARS_ROW,      "IO Period (yrs)",f.interest_only_years,    fmt='int',     note='0 = fully amortizing from day 1')
        input_row(ASM_CLOSING_COSTS_ROW, "Closing Costs",  f.closing_costs,          fmt='currency')

        # Calculated cells (formulas)
        row_eq = ASM_EQUITY_ROW
        cell_eq = ws.cell(row=row_eq, column=2, value=f"=B{ASM_PURCHASE_PRICE_ROW}-B{ASM_LOAN_AMOUNT_ROW}+B{ASM_CLOSING_COSTS_ROW}")
        cell_eq.font = self._font(bold=True, size=10)
        cell_eq.fill = self._fill(CALC_BG)
        self.fmt_currency(cell_eq)
        ws.cell(row=row_eq, column=1, value="Equity Invested").font = self._font(size=10)
        ws.cell(row=row_eq, column=4, value="Auto-calculated").font = self._font(size=9, italic=True, color='888888')

        row_ltv = ASM_LTV_ROW
        cell_ltv = ws.cell(row=row_ltv, column=2, value=f"=B{ASM_LOAN_AMOUNT_ROW}/B{ASM_PURCHASE_PRICE_ROW}")
        cell_ltv.font = self._font(bold=True, size=10)
        cell_ltv.fill = self._fill(CALC_BG)
        self.fmt_percent(cell_ltv, 1)
        ws.cell(row=row_ltv, column=1, value="LTV").font = self._font(size=10)
        ws.cell(row=row_ltv, column=4, value="Auto-calculated").font = self._font(size=9, italic=True, color='888888')

        # ── INCOME ────────────────────────────────────────────────────────
        section(19, "INCOME")
        input_row(20, "Annual Gross Rent (100% occ.)", i.gross_scheduled_rent, fmt='currency')
        input_row(ASM_GROSS_RENT_ROW,   "Gross Scheduled Rent",  i.gross_scheduled_rent, fmt='currency', note='Annual at 100% occupancy')
        input_row(ASM_VACANCY_ROW,      "Vacancy Rate",           i.vacancy_rate,          fmt='pct',      note='e.g. 0.05 = 5%')
        input_row(ASM_OTHER_INCOME_ROW, "Other Income",           i.other_income,          fmt='currency', note='Parking, laundry, etc. (annual)')
        input_row(ASM_RENT_GROWTH_ROW,  "Annual Rent Growth",     i.rent_growth_rate,      fmt='pct',      note='e.g. 0.03 = 3%/yr')

        # ── EXPENSES ─────────────────────────────────────────────────────
        section(25, "OPERATING EXPENSES")
        input_row(26, "Expense Growth Rate",        e.expense_growth_rate,      fmt='pct')
        input_row(ASM_TAXES_ROW,       "Property Taxes",          e.property_taxes,           fmt='currency', note='Annual')
        input_row(ASM_INSURANCE_ROW,   "Insurance",               e.insurance,                fmt='currency', note='Annual')
        input_row(ASM_MGMT_ROW,        "Property Management",     e.property_management,      fmt='pct' if e.property_management < 1 else 'currency', note='% of EGI (e.g. 0.08) or fixed $')
        input_row(ASM_MAINTENANCE_ROW, "Maintenance / Repairs",   e.maintenance_repairs,      fmt='currency', note='Annual')
        input_row(ASM_UTILITIES_ROW,   "Utilities",               e.utilities,                fmt='currency', note='Landlord-paid annual')
        input_row(ASM_LANDSCAPING_ROW, "Landscaping / Trash",     e.landscaping_trash,        fmt='currency', note='Annual')
        input_row(ASM_CAPEX_ROW,       "CapEx Reserve",           e.capex_reserve,            fmt='pct' if e.capex_reserve < 1 else 'currency', note='% of EGI (e.g. 0.05) or fixed $')
        input_row(ASM_EXP_GROWTH_ROW,  "Expense Growth Rate",     e.expense_growth_rate,      fmt='pct',      note='Annual escalation')

        # ── HOLD ASSUMPTIONS ─────────────────────────────────────────────
        section(35, "HOLD & EXIT ASSUMPTIONS")
        input_row(36, "",                           '',                         fmt='text')
        input_row(ASM_HOLD_YEARS_ROW,    "Hold Period (years)",    a.hold_period_years,        fmt='int',      note='Projection horizon')
        input_row(ASM_DISCOUNT_RATE_ROW, "Discount Rate (NPV)",    a.discount_rate,            fmt='pct',      note='e.g. 0.08 = 8%')
        input_row(ASM_EXIT_CAP_ROW,      "Exit Cap Rate",          a.exit_cap_rate,            fmt='pct',      note='Assumed sale cap rate')
        input_row(ASM_COST_OF_SALE_ROW,  "Cost of Sale",           a.cost_of_sale,             fmt='pct',      note='Broker + closing costs at exit')

        # ── WATERFALL (if enabled) ────────────────────────────────────────
        if wf.enabled:
            next_row = ASM_COST_OF_SALE_ROW + 2
            section(next_row, "EQUITY WATERFALL")
            input_row(next_row + 1, "LP Equity %",      wf.lp_equity_pct,   fmt='pct')
            input_row(next_row + 2, "GP Equity %",      wf.gp_equity_pct,   fmt='pct')
            input_row(next_row + 3, "Preferred Return", wf.preferred_return, fmt='pct')
            input_row(next_row + 4, "GP Promote %",     wf.promote_pct,     fmt='pct')

        # ── Legend ────────────────────────────────────────────────────────
        legend_row = ASM_COST_OF_SALE_ROW + (8 if wf.enabled else 3)
        ws.cell(row=legend_row, column=1, value="🟡 Yellow cells = editable inputs").font = self._font(size=9, italic=True)
        ws.cell(row=legend_row + 1, column=1, value="🟢 Green cells = auto-calculated (do not edit)").font = self._font(size=9, italic=True)
        ws.cell(row=legend_row + 2, column=1, value="Changes here flow through Annual Cash Flows, Summary, and Exit sheets.").font = self._font(size=9, italic=True, color='888888')
