from collections import defaultdict
from realty_model.renderers.base_renderer import BaseRenderer
from realty_model.models.results import FinancialResults, ExitScenario
from realty_model.models.property import ProFormaInput
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter


class ExitScenariosRenderer(BaseRenderer):
    def render(self, wb, results: FinancialResults, inp: ProFormaInput):
        ws = wb.create_sheet("Exit Scenarios")
        ws.sheet_view.showGridLines = False

        self.add_title_row(ws, 1, "Exit Scenario Analysis", 10)

        scenarios = results.exit_scenarios
        # Group by year, sub-group by cap rate
        years = sorted(set(s.year for s in scenarios))
        cap_rates = sorted(set(s.exit_cap_rate for s in scenarios))

        # Build lookup
        lookup = {(s.year, s.exit_cap_rate): s for s in scenarios}

        # Column layout: col 1 = metric label, cols 2.. = years
        n_data_cols = len(years)

        # Write cap rate headers at top
        row = 3
        metrics = [
            ("Exit Cap Rate",       'exit_cap_rate',      'percent2'),
            ("Gross Sale Price",    'gross_sale_price',   'currency'),
            ("Cost of Sale",        'cost_of_sale_dollars','currency'),
            ("Net Sale Proceeds",   'net_sale_proceeds',  'currency'),
            ("Loan Payoff",         'loan_payoff',        'currency'),
            ("Net Equity Proceeds", 'net_equity',         'currency'),
            ("Total Distributions", 'total_distributions','currency'),
            ("Total Profit",        'total_profit',       'currency'),
            ("Equity Multiple",     'equity_multiple',    'number'),
            ("Levered IRR",         'levered_irr',        'percent2'),
        ]

        # For each hold year, write a block
        col_offset = 2
        self.set_col_width(ws, 1, 22)

        # Write hold year headers
        hold_year_col = col_offset
        for y in years:
            count = len(cap_rates)
            end_col = hold_year_col + count - 1
            ws.merge_cells(start_row=row, start_column=hold_year_col,
                           end_row=row, end_column=end_col)
            cell = ws.cell(row=row, column=hold_year_col, value=f"Hold Year {y}")
            self.style_subheader(cell)
            for c in range(hold_year_col, end_col + 1):
                self.set_col_width(ws, c, 13)
            hold_year_col = end_col + 1

        row += 1

        # Write cap rate sub-headers
        cell_l = ws.cell(row=row, column=1, value="Exit Cap Rate →")
        self.style_header(cell_l)
        col = col_offset
        for y in years:
            for cap in cap_rates:
                cell = ws.cell(row=row, column=col, value=cap)
                self.style_header(cell)
                self.fmt_percent(cell, decimals=2)
                col += 1
        row += 1

        # Write metric rows
        for metric_label, field_name, fmt in metrics:
            is_irr = field_name == 'levered_irr'
            is_em = field_name == 'equity_multiple'

            cell_l = ws.cell(row=row, column=1, value=metric_label)
            bg = self.GRAY_BG if row % 2 == 0 else None
            cell_l.font = self._font(size=10)
            cell_l.alignment = self._align(h='left')
            if bg:
                cell_l.fill = self._fill(bg)

            col = col_offset
            for y in years:
                for cap in cap_rates:
                    s = lookup.get((y, cap))
                    val = getattr(s, field_name) if s else None
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = self._font(size=10)
                    cell.alignment = self._align(h='center')
                    if bg:
                        cell.fill = self._fill(bg)
                    if fmt == 'currency':
                        self.fmt_currency(cell)
                    elif fmt == 'percent2':
                        self.fmt_percent(cell, decimals=2)
                    elif fmt == 'number':
                        self.fmt_number(cell)
                    col += 1

            # Apply heat map to IRR and EM rows
            if is_irr and col > col_offset:
                range_str = (f"{get_column_letter(col_offset)}{row}:"
                             f"{get_column_letter(col - 1)}{row}")
                ws.conditional_formatting.add(range_str, ColorScaleRule(
                    start_type='num', start_value=0.05, start_color='C00000',
                    mid_type='num',   mid_value=0.12,   mid_color='FFFF00',
                    end_type='num',   end_value=0.20,   end_color='375623',
                ))

            row += 1
