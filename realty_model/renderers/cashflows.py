"""
Annual Cash Flows sheet — fully formula-driven.
All cells reference the Assumptions sheet so changing any
input cascades through rent growth, expenses, NOI, and cash flow.
"""
from openpyxl.utils import get_column_letter

from realty_model.renderers.base_renderer import BaseRenderer
from realty_model.renderers.constants import (
    ASM_EQUITY_ROW,
    ASM_GROSS_RENT_ROW, ASM_VACANCY_ROW, ASM_OTHER_INCOME_ROW, ASM_RENT_GROWTH_ROW,
    ASM_TAXES_ROW, ASM_INSURANCE_ROW, ASM_MGMT_ROW, ASM_MAINTENANCE_ROW,
    ASM_UTILITIES_ROW, ASM_LANDSCAPING_ROW, ASM_CAPEX_ROW, ASM_EXP_GROWTH_ROW,
    ASM_EXIT_CAP_ROW, ASM_COST_OF_SALE_ROW,
    CF_ROW_YEAR_HDR, CF_ROW_GROSS_RENT, CF_ROW_VACANCY, CF_ROW_OTHER_INCOME,
    CF_ROW_EGI, CF_ROW_TAXES, CF_ROW_INSURANCE, CF_ROW_MGMT,
    CF_ROW_MAINTENANCE, CF_ROW_UTILITIES, CF_ROW_LANDSCAPING, CF_ROW_CAPEX,
    CF_ROW_TOTAL_EXPENSES, CF_ROW_NOI, CF_ROW_DS_INTEREST, CF_ROW_DS_PRINCIPAL,
    CF_ROW_DS_TOTAL, CF_ROW_CASH_FLOW, CF_ROW_COC, CF_ROW_DSCR,
    CF_ROW_LOAN_BALANCE, CF_ROW_CUM_CF,
    CF_ROW_EXIT_VALUE, CF_ROW_SALE_COSTS, CF_ROW_NET_PROCEEDS,
    CF_ROW_NET_EQUITY, CF_ROW_TOTAL_RETURN, CF_ROW_EQUITY_MULTIPLE,
    CF_ROW_IRR_ARRAY,
    AMORT_SHEET, AMORT_COL_YEAR,
    AMORT_COL_INT, AMORT_COL_PRIN, AMORT_COL_BAL, AMORT_COL_MONTH,
    asm,
)
from realty_model.models.results import FinancialResults
from realty_model.models.property import ProFormaInput


def _col(year: int) -> int:
    """Year 1 → col 2 (B), Year 2 → col 3 (C), ..."""
    return year + 1


def _cl(year: int) -> str:
    """Column letter for given year (1-indexed)."""
    return get_column_letter(_col(year))


class AnnualCashFlowRenderer(BaseRenderer):

    def render(self, wb, results: FinancialResults, inp: ProFormaInput):
        ws = wb.create_sheet("Annual Cash Flows")
        ws.sheet_view.showGridLines = False

        hold = inp.assumptions.hold_period_years
        n_cols = hold + 1

        # Title
        self.add_title_row(ws, 1, "Annual Cash Flow Projections  —  Formulas update when Assumptions tab changes", n_cols)

        # Column widths
        self.set_col_width(ws, 1, 34)
        for y in range(1, hold + 1):
            self.set_col_width(ws, _col(y), 15)

        # Year header row
        ws.cell(row=CF_ROW_YEAR_HDR, column=1).fill = self._fill(self.HDR_BG)
        for y in range(1, hold + 1):
            cell = ws.cell(row=CF_ROW_YEAR_HDR, column=_col(y), value=y)
            self.style_header(cell)

        ws.freeze_panes = f'B{CF_ROW_YEAR_HDR + 1}'

        # Write labels and formulas
        self._write_labels(ws, hold)
        for y in range(1, hold + 1):
            self._write_year(ws, y, hold)

        # Hidden IRR array (used by Summary sheet)
        ws.row_dimensions[CF_ROW_IRR_ARRAY].hidden = True
        ws.cell(row=CF_ROW_IRR_ARRAY, column=1, value="IRR_ARRAY")
        # Col 2 = initial investment (negative)
        ws.cell(row=CF_ROW_IRR_ARRAY, column=2,
                value=f"=-{asm(ASM_EQUITY_ROW)}")
        for y in range(1, hold + 1):
            col = _col(y)
            cl = _cl(y)
            if y < hold:
                f = f"={cl}{CF_ROW_CASH_FLOW}"
            else:
                f = f"={cl}{CF_ROW_CASH_FLOW}+{cl}{CF_ROW_NET_EQUITY}"
            ws.cell(row=CF_ROW_IRR_ARRAY, column=col + 1, value=f)

    # ── Labels ────────────────────────────────────────────────────────────

    def _write_labels(self, ws, hold: int):
        def sec(row, text):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=hold + 1)
            self.style_section(ws.cell(row=row, column=1), text)

        def lbl(row, text, bold=False):
            c = ws.cell(row=row, column=1, value=text)
            c.font = self._font(bold=bold, size=10)
            c.alignment = self._align(h='left')
            if row % 2 == 0:
                c.fill = self._fill(self.GRAY_BG)

        sec(3,                     "INCOME")
        lbl(CF_ROW_GROSS_RENT,     "Gross Scheduled Rent")
        lbl(CF_ROW_VACANCY,        "(Less) Vacancy Loss")
        lbl(CF_ROW_OTHER_INCOME,   "Other Income")
        lbl(7,                     "")
        lbl(CF_ROW_EGI,            "Effective Gross Income (EGI)", bold=True)

        sec(9,                     "OPERATING EXPENSES")
        lbl(CF_ROW_TAXES,          "Property Taxes")
        lbl(CF_ROW_INSURANCE,      "Insurance")
        lbl(CF_ROW_MGMT,           "Property Management")
        lbl(CF_ROW_MAINTENANCE,    "Maintenance / Repairs")
        lbl(CF_ROW_UTILITIES,      "Utilities")
        lbl(CF_ROW_LANDSCAPING,    "Landscaping / Trash")
        lbl(CF_ROW_CAPEX,          "CapEx Reserve")
        lbl(17,                    "")
        lbl(CF_ROW_TOTAL_EXPENSES, "Total Operating Expenses", bold=True)

        lbl(19,                    "")
        lbl(CF_ROW_NOI,            "Net Operating Income (NOI)", bold=True)

        sec(21,                    "DEBT SERVICE  (sourced from Amortization Schedule tab)")
        lbl(CF_ROW_DS_INTEREST,    "Interest")
        lbl(CF_ROW_DS_PRINCIPAL,   "Principal")
        lbl(CF_ROW_DS_TOTAL,       "Total Debt Service", bold=True)

        lbl(25,                    "")
        lbl(CF_ROW_CASH_FLOW,      "Net Cash Flow (After Debt Service)", bold=True)
        lbl(CF_ROW_COC,            "Cash-on-Cash Return")
        lbl(CF_ROW_DSCR,           "Debt Service Coverage Ratio (DSCR)")

        sec(29,                    "EQUITY POSITION")
        lbl(CF_ROW_LOAN_BALANCE,   "Ending Loan Balance")
        lbl(CF_ROW_CUM_CF,         "Cumulative Net Cash Flow")

        sec(32,                    "EXIT ANALYSIS  (updates with Assumptions)")
        lbl(CF_ROW_EXIT_VALUE,     "Projected Exit Value  (NOI ÷ Exit Cap Rate)")
        lbl(CF_ROW_SALE_COSTS,     "Cost of Sale")
        lbl(CF_ROW_NET_PROCEEDS,   "Net Sale Proceeds")
        lbl(CF_ROW_NET_EQUITY,     "Net Equity at Exit  (proceeds − loan payoff)", bold=True)
        lbl(CF_ROW_TOTAL_RETURN,   "Total Return  (cum cash flow + net equity)", bold=True)
        lbl(CF_ROW_EQUITY_MULTIPLE,"Equity Multiple", bold=True)

    # ── Formulas for one year column ──────────────────────────────────────

    def _write_year(self, ws, y: int, hold: int):
        cl      = _cl(y)
        prev_cl = _cl(y - 1) if y > 1 else None
        col     = _col(y)

        def put(row, formula, fmt='currency', bold=False, bg=None):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = self._font(bold=bold, size=10)
            cell.alignment = self._align(h='right')
            default_bg = self.GRAY_BG if row % 2 == 0 else 'FFFFFF'
            cell.fill = self._fill(bg if bg else default_bg)
            if fmt == 'currency': self.fmt_currency(cell)
            elif fmt == 'pct':    self.fmt_percent(cell, 2)
            elif fmt == 'num':    self.fmt_number(cell)

        rg = asm(ASM_RENT_GROWTH_ROW)
        eg = asm(ASM_EXP_GROWTH_ROW)

        # ── Income ───────────────────────────────────────────────────────
        if y == 1:
            put(CF_ROW_GROSS_RENT,    f"={asm(ASM_GROSS_RENT_ROW)}")
            put(CF_ROW_OTHER_INCOME,  f"={asm(ASM_OTHER_INCOME_ROW)}")
        else:
            put(CF_ROW_GROSS_RENT,    f"={prev_cl}{CF_ROW_GROSS_RENT}*(1+{rg})")
            put(CF_ROW_OTHER_INCOME,  f"={prev_cl}{CF_ROW_OTHER_INCOME}*(1+{rg})")

        put(CF_ROW_VACANCY,       f"={cl}{CF_ROW_GROSS_RENT}*{asm(ASM_VACANCY_ROW)}")
        put(CF_ROW_EGI,
            f"={cl}{CF_ROW_GROSS_RENT}-{cl}{CF_ROW_VACANCY}+{cl}{CF_ROW_OTHER_INCOME}",
            bold=True, bg=self.DARK_GRAY)

        # ── Expenses ─────────────────────────────────────────────────────
        if y == 1:
            put(CF_ROW_TAXES,       f"={asm(ASM_TAXES_ROW)}")
            put(CF_ROW_INSURANCE,   f"={asm(ASM_INSURANCE_ROW)}")
            put(CF_ROW_MAINTENANCE, f"={asm(ASM_MAINTENANCE_ROW)}")
            put(CF_ROW_UTILITIES,   f"={asm(ASM_UTILITIES_ROW)}")
            put(CF_ROW_LANDSCAPING, f"={asm(ASM_LANDSCAPING_ROW)}")
        else:
            put(CF_ROW_TAXES,       f"={prev_cl}{CF_ROW_TAXES}*(1+{eg})")
            put(CF_ROW_INSURANCE,   f"={prev_cl}{CF_ROW_INSURANCE}*(1+{eg})")
            put(CF_ROW_MAINTENANCE, f"={prev_cl}{CF_ROW_MAINTENANCE}*(1+{eg})")
            put(CF_ROW_UTILITIES,   f"={prev_cl}{CF_ROW_UTILITIES}*(1+{eg})")
            put(CF_ROW_LANDSCAPING, f"={prev_cl}{CF_ROW_LANDSCAPING}*(1+{eg})")

        # Mgmt & CapEx: IF percent of EGI, ELSE treat as fixed $
        mgmt_ref  = asm(ASM_MGMT_ROW)
        capex_ref = asm(ASM_CAPEX_ROW)
        put(CF_ROW_MGMT,  f"=IF({mgmt_ref}<1,{cl}{CF_ROW_EGI}*{mgmt_ref},{mgmt_ref})")
        put(CF_ROW_CAPEX, f"=IF({capex_ref}<1,{cl}{CF_ROW_EGI}*{capex_ref},{capex_ref})")

        put(CF_ROW_TOTAL_EXPENSES,
            f"=SUM({cl}{CF_ROW_TAXES}:{cl}{CF_ROW_CAPEX})",
            bold=True, bg=self.DARK_GRAY)

        # ── NOI ──────────────────────────────────────────────────────────
        put(CF_ROW_NOI,
            f"={cl}{CF_ROW_EGI}-{cl}{CF_ROW_TOTAL_EXPENSES}",
            bold=True, bg=self.GREEN_BG)

        # ── Debt service (SUMIF against Amortization Schedule tab) ───────
        yr_range  = f"'Amortization Schedule'!${AMORT_COL_YEAR}:${AMORT_COL_YEAR}"
        int_range = f"'Amortization Schedule'!${AMORT_COL_INT}:${AMORT_COL_INT}"
        pri_range = f"'Amortization Schedule'!${AMORT_COL_PRIN}:${AMORT_COL_PRIN}"

        put(CF_ROW_DS_INTEREST,
            f"=SUMIF({yr_range},{cl}${CF_ROW_YEAR_HDR},{int_range})")
        put(CF_ROW_DS_PRINCIPAL,
            f"=SUMIF({yr_range},{cl}${CF_ROW_YEAR_HDR},{pri_range})")
        put(CF_ROW_DS_TOTAL,
            f"={cl}{CF_ROW_DS_INTEREST}+{cl}{CF_ROW_DS_PRINCIPAL}",
            bold=True, bg=self.DARK_GRAY)

        # ── Cash flow ────────────────────────────────────────────────────
        put(CF_ROW_CASH_FLOW,
            f"={cl}{CF_ROW_NOI}-{cl}{CF_ROW_DS_TOTAL}",
            bold=True, bg=self.GREEN_BG)
        put(CF_ROW_COC,
            f"=IFERROR({cl}{CF_ROW_CASH_FLOW}/{asm(ASM_EQUITY_ROW)},0)",
            fmt='pct')
        put(CF_ROW_DSCR,
            f"=IFERROR({cl}{CF_ROW_NOI}/{cl}{CF_ROW_DS_TOTAL},0)",
            fmt='num')

        # ── Equity ───────────────────────────────────────────────────────
        # Ending loan balance: VLOOKUP on month = year*12 in Amortization tab
        bal_range = f"'Amortization Schedule'!${AMORT_COL_MONTH}:${AMORT_COL_BAL}"
        put(CF_ROW_LOAN_BALANCE,
            f"=IFERROR(VLOOKUP({cl}${CF_ROW_YEAR_HDR}*12,{bal_range},6,FALSE),0)")

        if y == 1:
            put(CF_ROW_CUM_CF, f"={cl}{CF_ROW_CASH_FLOW}")
        else:
            put(CF_ROW_CUM_CF, f"={prev_cl}{CF_ROW_CUM_CF}+{cl}{CF_ROW_CASH_FLOW}")

        # ── Exit analysis ─────────────────────────────────────────────────
        exit_cap = asm(ASM_EXIT_CAP_ROW)
        cos_ref  = asm(ASM_COST_OF_SALE_ROW)

        put(CF_ROW_EXIT_VALUE,
            f"=IFERROR({cl}{CF_ROW_NOI}/{exit_cap},0)")
        put(CF_ROW_SALE_COSTS,
            f"={cl}{CF_ROW_EXIT_VALUE}*{cos_ref}")
        put(CF_ROW_NET_PROCEEDS,
            f"={cl}{CF_ROW_EXIT_VALUE}-{cl}{CF_ROW_SALE_COSTS}")
        put(CF_ROW_NET_EQUITY,
            f"={cl}{CF_ROW_NET_PROCEEDS}-{cl}{CF_ROW_LOAN_BALANCE}",
            bold=True)
        put(CF_ROW_TOTAL_RETURN,
            f"={cl}{CF_ROW_CUM_CF}+{cl}{CF_ROW_NET_EQUITY}",
            bold=True, bg=self.GREEN_BG)
        put(CF_ROW_EQUITY_MULTIPLE,
            f"=IFERROR({cl}{CF_ROW_TOTAL_RETURN}/{asm(ASM_EQUITY_ROW)},0)",
            fmt='num', bold=True)
