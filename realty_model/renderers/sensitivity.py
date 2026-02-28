from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from realty_model.renderers.base_renderer import BaseRenderer
from realty_model.models.results import FinancialResults
from realty_model.models.property import ProFormaInput


class SensitivityRenderer(BaseRenderer):
    def render(self, wb, results: FinancialResults, inp: ProFormaInput):
        ws = wb.create_sheet("Sensitivity Analysis")
        ws.sheet_view.showGridLines = False

        self.add_title_row(ws, 1, "Sensitivity Analysis", 12)

        self._write_table(
            ws,
            title="Levered IRR  (Exit Cap Rate  ×  Vacancy Rate)",
            df=results.sensitivity_irr,
            start_row=3,
            fmt='percent2',
            heatmap_low=0.05,
            heatmap_mid=0.10,
            heatmap_high=0.20,
            heatmap_low_color='C00000',
            heatmap_mid_color='FFFF00',
            heatmap_high_color='375623',
        )

        n_irr_rows = len(results.sensitivity_irr) + 4  # title + header + data + spacer
        coc_start = 3 + n_irr_rows + 2

        self._write_table(
            ws,
            title="Cash-on-Cash Return (Year 1)  ×  Vacancy Rate",
            df=results.sensitivity_coc,
            start_row=coc_start,
            fmt='percent2',
            heatmap_low=0.02,
            heatmap_mid=0.06,
            heatmap_high=0.12,
            heatmap_low_color='C00000',
            heatmap_mid_color='FFFF00',
            heatmap_high_color='375623',
        )

    def _write_table(self, ws, title, df, start_row, fmt,
                     heatmap_low, heatmap_mid, heatmap_high,
                     heatmap_low_color, heatmap_mid_color, heatmap_high_color):
        row = start_row
        n_cols = len(df.columns)
        total_cols = n_cols + 1  # label + data cols

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
        self.style_subheader(ws.cell(row=row, column=1), title)
        row += 1

        # Row index label
        cell = ws.cell(row=row, column=1, value=df.index.name or "")
        self.style_header(cell)
        self.set_col_width(ws, 1, 18)

        # Column headers
        for col_idx, col_name in enumerate(df.columns, 2):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            self.style_header(cell)
            self.set_col_width(ws, col_idx, 13)
        row += 1

        data_start_row = row
        for i, (index_val, series) in enumerate(df.iterrows()):
            bg = self.GRAY_BG if row % 2 == 0 else None
            cell_l = ws.cell(row=row, column=1, value=index_val)
            cell_l.font = self._font(bold=True, size=10)
            cell_l.alignment = self._align(h='center')
            if bg:
                cell_l.fill = self._fill(bg)

            for col_idx, val in enumerate(series, 2):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = self._font(size=10)
                cell.alignment = self._align(h='center')
                if fmt == 'percent2':
                    self.fmt_percent(cell, decimals=2)
                elif fmt == 'currency':
                    self.fmt_currency(cell)
            row += 1

        data_end_row = row - 1
        # Apply 3-color heat map to data cells
        col_start_letter = get_column_letter(2)
        col_end_letter = get_column_letter(n_cols + 1)
        range_str = f"{col_start_letter}{data_start_row}:{col_end_letter}{data_end_row}"

        rule = ColorScaleRule(
            start_type='num', start_value=heatmap_low, start_color=heatmap_low_color,
            mid_type='num',   mid_value=heatmap_mid,   mid_color=heatmap_mid_color,
            end_type='num',   end_value=heatmap_high,  end_color=heatmap_high_color,
        )
        ws.conditional_formatting.add(range_str, rule)
